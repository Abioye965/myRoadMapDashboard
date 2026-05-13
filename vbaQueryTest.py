from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter

input_file = "/mnt/data/Sample Auto.xlsx"

wb = load_workbook(input_file)

roadmap = wb["RoadMap"]
requests = wb["Requests"]

last_row = roadmap.max_row

# Create table
table_name = "RoadmapTable"
if table_name not in roadmap.tables:
    tbl = Table(displayName=table_name, ref=f"A1:D{last_row}")
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tbl.tableStyleInfo = style
    roadmap.add_table(tbl)

# Helper sheet
if "Helper" in wb.sheetnames:
    helper = wb["Helper"]
else:
    helper = wb.create_sheet("Helper")

helper.sheet_state = "hidden"

categories = {}

for r in range(2, last_row + 1):
    cat = roadmap[f"B{r}"].value
    item = roadmap[f"C{r}"].value

    if cat and item:
        categories.setdefault(cat, []).append(item)

# Categories list
helper["A1"] = "Categories"

for i, cat in enumerate(sorted(categories.keys()), start=2):
    helper[f"A{i}"] = cat

# Named range for categories
cat_end = len(categories) + 1
wb.defined_names.add(
    DefinedName("CategoryList", attr_text=f"Helper!$A$2:$A${cat_end}")
)

# Create named ranges for each category
col_num = 2

for cat, items in categories.items():
    safe_name = cat.replace(" ", "").replace("-", "")

    helper.cell(row=1, column=col_num).value = safe_name

    unique_items = sorted(set(items))

    for idx, item in enumerate(unique_items, start=2):
        helper.cell(row=idx, column=col_num).value = item

    col_letter = get_column_letter(col_num)

    rng = f"Helper!${col_letter}$2:${col_letter}${len(unique_items)+1}"

    wb.defined_names.add(
        DefinedName(safe_name, attr_text=rng)
    )

    col_num += 1

# Validation for Category J
dv_cat = DataValidation(type="list", formula1="=CategoryList")
requests.add_data_validation(dv_cat)
dv_cat.add("J2:J1000")

# Dependent validation for Work Item K
dv_work = DataValidation(
    type="list",
    formula1='=INDIRECT(SUBSTITUTE(SUBSTITUTE($J2," ",""),"-",""))'
)

requests.add_data_validation(dv_work)
dv_work.add("K2:K1000")

# Auto lookup formula for Code L
for row in range(2, 1001):
    requests[f"L{row}"] = (
        f'=IFERROR(INDEX(RoadMap!$D:$D,MATCH(K{row},RoadMap!$C:$C,0)),"")'
    )

output_file = "/mnt/data/Sample_Auto_Completed.xlsx"

wb.save(output_file)

print(output_file)