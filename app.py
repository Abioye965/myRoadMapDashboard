import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Auto Roadmap Dashboard", page_icon="🗺️", layout="wide")

st.markdown("""
<style>
    .block-container { padding-top: 1.5rem; }
    .stTabs [data-baseweb="tab"] { font-size: 15px; font-weight: 600; }
    .stat-card {
        background: #f8f9fb; border-radius: 10px;
        padding: 16px 20px; border-left: 4px solid #4f6ef7; margin-bottom: 8px;
    }
    .stat-label { color: #666; font-size: 13px; margin-bottom: 2px; }
    .stat-value { font-size: 22px; font-weight: 700; color: #1a1a2e; }
    .section-title { font-size: 17px; font-weight: 700; margin: 18px 0 8px 0; }
</style>
""", unsafe_allow_html=True)

st.title("🗺️ Auto Roadmap Dashboard")
st.caption("Upload your workbook, browse the Roadmap, fill Requests, and export a completed Excel file.")

uploaded = st.file_uploader("**Upload Sample Auto.xlsx**", type=["xlsx"])
if not uploaded:
    st.info("👆 Upload your Excel file above to get started.")
    st.stop()


@st.cache_data
def load_data(file_bytes: bytes):
    wb = load_workbook(io.BytesIO(file_bytes))

    # ── RoadMap: Row 1 is blank, Row 2 = headers, Row 3+ = data ──
    roadmap_ws = wb["RoadMap"]
    raw_headers = [roadmap_ws.cell(2, c).value for c in range(1, 5)]

    seen = {}
    headers = []
    for h in raw_headers:
        h = str(h).strip() if h is not None else "Column"
        count = seen.get(h, 0)
        headers.append(f"{h}_{count}" if count > 0 else h)
        seen[h] = count + 1

    rows = []
    last_data_row = 2
    for r in range(3, roadmap_ws.max_row + 1):
        row = [roadmap_ws.cell(r, c).value for c in range(1, 5)]
        if row[1] is None and row[2] is None:
            continue
        rows.append([str(v).strip() if v is not None else "" for v in row])
        last_data_row = r

    df_roadmap = pd.DataFrame(rows, columns=headers)

    categories: dict = {}
    for _, row in df_roadmap.iterrows():
        cat, item, code = row.iloc[1], row.iloc[2], row.iloc[3]
        if cat and item:
            categories.setdefault(cat, {})[item] = code

    # ── Requests: Row 1 = headers, Row 2+ = data ──
    req_ws = wb["Requests"]
    raw_req_h = [req_ws.cell(1, c).value or f"Col{c}" for c in range(1, 13)]
    seen2 = {}
    safe_req_h = []
    for h in raw_req_h:
        h = str(h).strip()
        count = seen2.get(h, 0)
        safe_req_h.append(f"{h}_{count}" if count > 0 else h)
        seen2[h] = count + 1

    existing = []
    for r in range(2, req_ws.max_row + 1):
        row = [req_ws.cell(r, c).value for c in range(1, 13)]
        if any(v is not None for v in row):
            existing.append([str(v).strip() if v is not None else "" for v in row])

    df_req = pd.DataFrame(existing, columns=safe_req_h) if existing else pd.DataFrame(columns=safe_req_h)
    return df_roadmap, categories, df_req, safe_req_h, last_data_row


file_bytes = uploaded.read()
df_roadmap, categories, df_existing_requests, req_headers, last_data_row = load_data(file_bytes)

if "new_requests" not in st.session_state:
    st.session_state.new_requests = []

tab_roadmap, tab_form, tab_export = st.tabs([
    "📋  RoadMap Viewer", "📝  Requests Form", "📤  Export Excel"
])


# ── TAB 1: ROADMAP VIEWER ─────────────────────────────────────────────────────
with tab_roadmap:
    ca, cb, cc = st.columns(3)
    with ca:
        st.markdown(f'<div class="stat-card"><div class="stat-label">Total Items</div><div class="stat-value">{len(df_roadmap)}</div></div>', unsafe_allow_html=True)
    with cb:
        st.markdown(f'<div class="stat-card"><div class="stat-label">Categories</div><div class="stat-value">{len(categories)}</div></div>', unsafe_allow_html=True)
    with cc:
        st.markdown(f'<div class="stat-card"><div class="stat-label">Work Items</div><div class="stat-value">{sum(len(v) for v in categories.values())}</div></div>', unsafe_allow_html=True)

    st.markdown("---")

    cs1, cs2, cs3 = st.columns([2, 1, 1])
    with cs1:
        search = st.text_input("🔍 Search", placeholder="Type to filter any column…")
    with cs2:
        cat_filter = st.selectbox("Filter by Category", ["All"] + sorted(categories.keys()))
    with cs3:
        if cat_filter != "All":
            wi_opts = ["All"] + sorted(categories.get(cat_filter, {}).keys())
        else:
            wi_opts = ["All"] + sorted({k for v in categories.values() for k in v})
        wi_filter = st.selectbox("Filter by Work Item", wi_opts)

    cat_col  = df_roadmap.columns[1]
    item_col = df_roadmap.columns[2]
    filtered = df_roadmap.copy()

    if cat_filter != "All":
        filtered = filtered[filtered[cat_col] == cat_filter]
    if wi_filter != "All":
        filtered = filtered[filtered[item_col] == wi_filter]
    if search:
        mask = filtered.apply(
            lambda row: row.astype(str).str.contains(search, case=False, na=False).any(), axis=1
        )
        filtered = filtered[mask]

    st.dataframe(filtered, use_container_width=True, height=420)
    st.caption(f"Showing **{len(filtered)}** of **{len(df_roadmap)}** rows")


# ── TAB 2: REQUESTS FORM ──────────────────────────────────────────────────────
with tab_form:
    st.markdown('<div class="section-title">Add a New Request</div>', unsafe_allow_html=True)

    with st.container(border=True):
        c1, c2 = st.columns([1, 2])
        with c1:
            date_val = st.text_input("DATE (dd/mm/yy)")
        with c2:
            purpose_val = st.text_area("PURPOSE", height=80)

        c3, c4, c5, c6 = st.columns(4)
        with c3:
            qty_val = st.text_input("QTY")
        with c4:
            unit_val = st.text_input("UNIT")
        with c5:
            rate_val = st.text_input("RATE (₦)")
        with c6:
            total_val = st.text_input("TOTAL COST (₦)")

        c7, c8 = st.columns(2)
        with c7:
            estate_val = st.text_input("ESTATE")
        with c8:
            plot_val = st.text_input("PLOT NUMBER")

        c9, c10, c11 = st.columns(3)
        with c9:
            selected_cat = st.selectbox("CATEGORY (STAGES)", sorted(categories.keys()))
        with c10:
            work_items = sorted(categories.get(selected_cat, {}).keys())
            selected_item = st.selectbox("WORK ITEM / TASK SUB-STAGE", work_items)
        with c11:
            auto_code = categories.get(selected_cat, {}).get(selected_item, "")
            st.text_input("CODE (auto-filled)", value=auto_code, disabled=True)

        if st.button("➕ Add Request", use_container_width=True, type="primary"):
            next_sn = len(df_existing_requests) + len(st.session_state.new_requests) + 1
            st.session_state.new_requests.append({
                req_headers[0]: next_sn,
                req_headers[1]: date_val,
                req_headers[2]: purpose_val,
                req_headers[3]: qty_val,
                req_headers[4]: unit_val,
                req_headers[5]: rate_val,
                req_headers[6]: total_val,
                req_headers[7]: estate_val,
                req_headers[8]: plot_val,
                req_headers[9]: selected_cat,
                req_headers[10]: selected_item,
                req_headers[11]: auto_code,
            })
            st.success(f"Added: **{selected_cat}** → **{selected_item}** → `{auto_code}`")

    if not df_existing_requests.empty:
        st.markdown('<div class="section-title">Existing Requests (from file)</div>', unsafe_allow_html=True)
        st.dataframe(df_existing_requests, use_container_width=True)

    st.markdown('<div class="section-title">New Requests (this session)</div>', unsafe_allow_html=True)
    if st.session_state.new_requests:
        st.dataframe(pd.DataFrame(st.session_state.new_requests), use_container_width=True)

        b1, b2, b3 = st.columns([1, 1, 4])
        with b1:
            del_idx = st.number_input("Row # to delete", min_value=1,
                                      max_value=len(st.session_state.new_requests), step=1)
        with b2:
            st.write("")
            st.write("")
            if st.button("🗑 Delete Row"):
                st.session_state.new_requests.pop(int(del_idx) - 1)
                st.rerun()
        with b3:
            st.write("")
            st.write("")
            if st.button("🧹 Clear All"):
                st.session_state.new_requests = []
                st.rerun()
    else:
        st.info("No new requests added yet.")


# ── TAB 3: EXPORT ─────────────────────────────────────────────────────────────
with tab_export:
    st.markdown('<div class="section-title">Export Summary</div>', unsafe_allow_html=True)

    ce1, ce2, ce3 = st.columns(3)
    with ce1:
        st.markdown(f'<div class="stat-card"><div class="stat-label">Existing Requests</div><div class="stat-value">{len(df_existing_requests)}</div></div>', unsafe_allow_html=True)
    with ce2:
        st.markdown(f'<div class="stat-card"><div class="stat-label">New (this session)</div><div class="stat-value">{len(st.session_state.new_requests)}</div></div>', unsafe_allow_html=True)
    with ce3:
        total_r = len(df_existing_requests) + len(st.session_state.new_requests)
        st.markdown(f'<div class="stat-card"><div class="stat-label">Total After Export</div><div class="stat-value">{total_r}</div></div>', unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("""
    The exported file will include:
    - ✅ **RoadmapTable** — formatted Excel table on the RoadMap sheet
    - ✅ **Helper sheet** — hidden sheet with named ranges per category
    - ✅ **Data Validation** — Category (J) and Work Item (K) dropdowns
    - ✅ **Auto Code formula** — `IFERROR(INDEX/MATCH)` in column L
    - ✅ **Your new requests** — written into the Requests sheet
    """)

    if st.button("⚙️ Generate & Download Excel", type="primary", use_container_width=True):
        with st.spinner("Building your Excel file…"):
            wb_out = load_workbook(io.BytesIO(file_bytes))
            rm_ws  = wb_out["RoadMap"]
            rq_ws  = wb_out["Requests"]

            # RoadMap table (header on row 2, data from row 3)
            tname = "RoadmapTable"
            if tname not in rm_ws.tables:
                tbl = Table(displayName=tname, ref=f"A2:D{last_data_row}")
                tbl.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False, showLastColumn=False,
                    showRowStripes=True, showColumnStripes=False,
                )
                rm_ws.add_table(tbl)

            # Helper sheet
            helper = wb_out["Helper"] if "Helper" in wb_out.sheetnames else wb_out.create_sheet("Helper")
            helper.sheet_state = "hidden"

            cats: dict = {}
            for r in range(3, rm_ws.max_row + 1):
                cat  = rm_ws.cell(r, 2).value
                item = rm_ws.cell(r, 3).value
                if cat and item:
                    cats.setdefault(str(cat), []).append(str(item))

            helper["A1"] = "Categories"
            for i, cat in enumerate(sorted(cats.keys()), start=2):
                helper[f"A{i}"] = cat

            wb_out.defined_names.add(
                DefinedName("CategoryList", attr_text=f"Helper!$A$2:$A${len(cats)+1}")
            )

            col_num = 2
            for cat, items in cats.items():
                safe = cat.replace(" ", "").replace("-", "").replace("/", "")
                helper.cell(row=1, column=col_num).value = safe
                unique = sorted(set(items))
                for idx, item in enumerate(unique, start=2):
                    helper.cell(row=idx, column=col_num).value = item
                cl = get_column_letter(col_num)
                wb_out.defined_names.add(
                    DefinedName(safe, attr_text=f"Helper!${cl}$2:${cl}${len(unique)+1}")
                )
                col_num += 1

            # Write new requests — find first empty row via col B (DATE)
            next_row = 2
            while rq_ws.cell(next_row, 2).value is not None:
                next_row += 1

            for req in st.session_state.new_requests:
                for ci, val in enumerate(req.values(), start=1):
                    rq_ws.cell(next_row, ci).value = val
                next_row += 1

            # Validations
            dv_cat = DataValidation(type="list", formula1="=CategoryList")
            rq_ws.add_data_validation(dv_cat)
            dv_cat.add("J2:J1000")

            dv_work = DataValidation(
                type="list",
                formula1='=INDIRECT(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE($J2," ",""),"-",""),"/",""))',
            )
            rq_ws.add_data_validation(dv_work)
            dv_work.add("K2:K1000")

            for row in range(2, 1001):
                rq_ws[f"L{row}"] = f'=IFERROR(INDEX(RoadMap!$D:$D,MATCH(K{row},RoadMap!$C:$C,0)),"")'

            output = io.BytesIO()
            wb_out.save(output)
            output.seek(0)

        st.success("✅ File ready!")
        st.download_button(
            label="📥 Download Sample_Auto_Completed.xlsx",
            data=output,
            file_name="Sample_Auto_Completed.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
